#!/usr/bin/env python
# coding: utf-8

# In[2]:


from openpyxl import load_workbook

workbook = load_workbook(filename = r"C:\Users\easalazarm\Documents\IDEAM\IDEAM\Turnos\Diciembre\Tablas\Salidas\temperaturamax.xlsx")
boletin = load_workbook(filename = r"C:\Users\easalazarm\Documents\IDEAM\IDEAM\Turnos\Diciembre\Tablas\Boletin\boletin_26_dic_2023.xlsx" )
workbook.sheetnames
#boletin.sheetnames


# In[3]:


hoja = workbook['max']
hoja["D5"].value


# In[4]:


hoja2 = boletin["DIARIO"]
hoja3 = boletin["ENVIAR"]
hoja2["FU4"].value


# In[5]:


i = hoja2.cell(row=4, column= 177).value
i


# In[15]:


#hoja.cell(row = 5, column = 80 - (31 - i)).value = hoja2.cell(row = 9, column = 177).value


# In[17]:

columna_tmax_i = 39 #80  #49 + 31 #39

def llenar_datos(hoja1, hoja2, i, columna_tmax):
    
    #San Andrés
    hoja1.cell(row = 5, column = columna_tmax - (31 - i)).value = hoja2.cell(row = 9, column = 177).value
    #Providencia
    hoja1.cell(row = 6, column = columna_tmax - (31 - i)).value = hoja2.cell(row = 10, column = 177).value
    #Santa mayta
    hoja1.cell(row = 7, column = columna_tmax - (31 - i)).value = hoja2.cell(row = 11, column = 177).value
    #Cartagena
    hoja1.cell(row = 8, column = columna_tmax - (31 - i)).value = hoja2.cell(row = 12, column = 177).value
    #Barranquilla
    hoja1.cell(row = 9, column = columna_tmax - (31 - i)).value = hoja2.cell(row = 13, column = 177).value
    #Riohacha 
    hoja1.cell(row = 10, column = columna_tmax - (31 - i)).value = hoja2.cell(row = 14, column = 177).value
    #Valledupar
    hoja1.cell(row = 11, column = columna_tmax - (31 - i)).value = hoja2.cell(row = 15, column = 177).value
    #Monteria
    hoja1.cell(row = 13, column = columna_tmax - (31 - i)).value = hoja2.cell(row = 16, column = 177).value
    #Apartado
    hoja1.cell(row = 14, column = columna_tmax - (31 - i)).value = hoja2.cell(row = 20, column = 177).value
    #Barrancabermeja
    hoja1.cell(row = 16, column = columna_tmax - (31 - i)).value = hoja2.cell(row = 17, column = 177).value
    #Bucaramanga
    hoja1.cell(row = 17, column = columna_tmax - (31 - i)).value = hoja2.cell(row = 18, column = 177).value
    #Cucuta
    hoja1.cell(row = 18, column = columna_tmax - (31 - i)).value = hoja2.cell(row = 19, column = 177).value
    #Medellin
    hoja1.cell(row = 19, column = columna_tmax - (31 - i)).value = hoja2.cell(row = 21, column = 177).value
    #Rionegro
    hoja1.cell(row = 20, column = columna_tmax - (31 - i)).value = hoja2.cell(row = 22, column = 177).value
    #Pereira
    hoja1.cell(row = 22, column = columna_tmax - (31 - i)).value = hoja2.cell(row = 23, column = 177).value
    #Armenia
    hoja1.cell(row = 23, column = columna_tmax - (31 - i)).value = hoja2.cell(row = 24, column = 177).value
    #Ibague
    hoja1.cell(row = 24, column = columna_tmax - (31 - i)).value = hoja2.cell(row = 25, column = 177).value
    #Bogota
    hoja1.cell(row = 25, column = columna_tmax - (31 - i)).value = hoja2.cell(row = 26, column = 177).value
    #Cali
    hoja1.cell(row = 28, column = columna_tmax - (31 - i)).value = hoja2.cell(row = 27, column = 177).value
    #Neiva
    hoja1.cell(row = 30, column = columna_tmax - (31 - i)).value = hoja2.cell(row = 28, column = 177).value
    #Chachagui
    hoja1.cell(row = 31, column = columna_tmax - (31 - i)).value = hoja2.cell(row = 29, column = 177).value
    #Ipiales
    hoja1.cell(row = 32, column = columna_tmax - (31 - i)).value = hoja2.cell(row = 30, column = 177).value
    #Quibdo
    hoja1.cell(row = 35, column = columna_tmax - (31 - i)).value = hoja2.cell(row = 31, column = 177).value
    #Arauca
    hoja1.cell(row = 39, column = columna_tmax - (31 - i)).value = hoja2.cell(row = 32, column = 177).value
    #Puerto Carreño
    hoja1.cell(row = 40, column = columna_tmax - (31 - i)).value = hoja2.cell(row = 33, column = 177).value
    #Villavicencio
    hoja1.cell(row = 41, column = columna_tmax - (31 - i)).value = hoja2.cell(row = 34, column = 177).value
    #Leticia
    hoja1.cell(row = 43, column = columna_tmax - (31 - i)).value = hoja2.cell(row = 35, column = 177).value
    
    ############ Areas operativas ########
    ####
    
#    j = 0
#    
#    for r in range(1, hoja1.max_row+1):
#        for c in range(1, hoja1.max_column+1):
#            s = str(hoja1.cell(r,c).value)
#            if s == "/" in s: 
#                hoja1.cell(r,c).value = None 
#                #print("row {} col {} : {}".format(r,c,s))
#                j += 1
#    
    return workbook.save(filename =   r"C:\Users\easalazarm\Documents\IDEAM\IDEAM\Turnos\Diciembre\Tablas\Salidas\temperaturamax.xlsx")    


funcion = llenar_datos(hoja, hoja2, i, columna_tmax_i)
funcion


# In[ ]:




