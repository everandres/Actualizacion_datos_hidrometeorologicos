#!/usr/bin/env python
# coding: utf-8

# In[3]:


from openpyxl import load_workbook

workbook = load_workbook(filename = r"C:\Users\easalazarm\Documents\IDEAM\IDEAM\Turnos\Diciembre\Tablas\Salidas\temperaturamed.xlsx")
boletin = load_workbook(filename = r"C:\Users\easalazarm\Documents\IDEAM\IDEAM\Turnos\Diciembre\Tablas\Boletin\boletin_26_dic_2023.xlsx" )
workbook.sheetnames
#boletin.sheetnames


# In[4]:


hoja = workbook['med']
hoja["D5"].value


# In[5]:


hoja2 = boletin["DIARIO"]
hoja3 = boletin["ENVIAR"]
hoja2["FU4"].value


# In[6]:


i = hoja2.cell(row=4, column= 177).value
i


# In[23]:


#hoja.cell(row = 5, column = 71 - (31 - i)).value = hoja2.cell(row = 9, column = 175).value
#hoja.cell(row = 5, column = 71 - (31 - i)).value


# In[24]:

columna_tmed_i = 39 #71  #40 + 31 #39

def llenar_datos(hoja1, hoja2, i, columna_tmed):
    
    #San Andrés
    hoja1.cell(row = 5, column = columna_tmed - (31 - i)).value = hoja2.cell(row = 9, column = 175).value
    #Providencia
    hoja1.cell(row = 6, column = columna_tmed - (31 - i)).value = hoja2.cell(row = 10, column = 175).value
    #Santa abrta
    hoja1.cell(row = 7, column = columna_tmed - (31 - i)).value = hoja2.cell(row = 11, column = 175).value
    #Cartagena
    hoja1.cell(row = 8, column = columna_tmed - (31 - i)).value = hoja2.cell(row = 12, column = 175).value
    #Barranquilla
    hoja1.cell(row = 9, column = columna_tmed - (31 - i)).value = hoja2.cell(row = 13, column = 175).value
    #Riohacha 
    hoja1.cell(row = 10, column = columna_tmed - (31 - i)).value = hoja2.cell(row = 14, column = 175).value
    #Valledupar
    hoja1.cell(row = 11, column = columna_tmed - (31 - i)).value = hoja2.cell(row = 15, column = 175).value
    #Monteria
    hoja1.cell(row = 13, column = columna_tmed - (31 - i)).value = hoja2.cell(row = 16, column = 175).value
    #Apartado
    hoja1.cell(row = 14, column = columna_tmed - (31 - i)).value = hoja2.cell(row = 20, column = 175).value
    #Barrancabermeja
    hoja1.cell(row = 16, column = columna_tmed - (31 - i)).value = hoja2.cell(row = 17, column = 175).value
    #Bucaramanga
    hoja1.cell(row = 17, column = columna_tmed - (31 - i)).value = hoja2.cell(row = 18, column = 175).value
    #Cucuta
    hoja1.cell(row = 18, column = columna_tmed - (31 - i)).value = hoja2.cell(row = 19, column = 175).value
    #Medellin
    hoja1.cell(row = 19, column = columna_tmed - (31 - i)).value = hoja2.cell(row = 21, column = 175).value
    #Rionegro
    hoja1.cell(row = 20, column = columna_tmed - (31 - i)).value = hoja2.cell(row = 22, column = 175).value
    #Pereira
    hoja1.cell(row = 22, column = columna_tmed - (31 - i)).value = hoja2.cell(row = 23, column = 175).value
    #Armenia
    hoja1.cell(row = 23, column = columna_tmed - (31 - i)).value = hoja2.cell(row = 24, column = 175).value
    #Ibague
    hoja1.cell(row = 24, column = columna_tmed - (31 - i)).value = hoja2.cell(row = 25, column = 175).value
    #Bogota
    hoja1.cell(row = 25, column = columna_tmed - (31 - i)).value = hoja2.cell(row = 26, column = 175).value
    #Cali
    hoja1.cell(row = 28, column = columna_tmed - (31 - i)).value = hoja2.cell(row = 27, column = 175).value
    #Neiva
    hoja1.cell(row = 30, column = columna_tmed - (31 - i)).value = hoja2.cell(row = 28, column = 175).value
    #Chachagui
    hoja1.cell(row = 31, column = columna_tmed - (31 - i)).value = hoja2.cell(row = 29, column = 175).value
    #Ipiales
    hoja1.cell(row = 32, column = columna_tmed - (31 - i)).value = hoja2.cell(row = 30, column = 175).value
    #Quibdo
    hoja1.cell(row = 35, column = columna_tmed - (31 - i)).value = hoja2.cell(row = 31, column = 175).value
    #Arauca
    hoja1.cell(row = 39, column = columna_tmed - (31 - i)).value = hoja2.cell(row = 32, column = 175).value
    #Puerto Carreño
    hoja1.cell(row = 40, column = columna_tmed - (31 - i)).value = hoja2.cell(row = 33, column = 175).value
    #Villavicencio
    hoja1.cell(row = 41, column = columna_tmed - (31 - i)).value = hoja2.cell(row = 34, column = 175).value
    #Leticia
    hoja1.cell(row = 43, column = columna_tmed - (31 - i)).value = hoja2.cell(row = 35, column = 175).value
    
    ############ Areas operativas ########
    
    #########
    
#    j = 0
#    
#    for r in range(1, hoja1.max_row+1):
#        for c in range(1, hoja1.max_column+1):
#            s = str(hoja1.cell(r,c).value)
#            if s == "/" in s: 
#                hoja1.cell(r,c).value = None 
#                #print("row {} col {} : {}".format(r,c,s))
#                j += 1
    
    return workbook.save(filename = r"C:\Users\easalazarm\Documents\IDEAM\IDEAM\Turnos\Diciembre\Tablas\Salidas\temperaturamed.xlsx")    


funcion = llenar_datos(hoja, hoja2, i, columna_tmed_i)
funcion


# In[ ]:




