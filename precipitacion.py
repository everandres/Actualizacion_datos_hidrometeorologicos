#!/usr/bin/env python
# coding: utf-8

# In[1]:


from openpyxl import load_workbook

workbook = load_workbook(filename = r"C:\Users\easalazarm\Documents\IDEAM\IDEAM\Turnos\Diciembre\Tablas\Salidas\precipitacion.xlsx")
boletin = load_workbook(filename = r"C:\Users\easalazarm\Documents\IDEAM\IDEAM\Turnos\Diciembre\Tablas\Boletin\boletin_26_dic_2023.xlsx" )
workbook.sheetnames
#boletin.sheetnames


# In[2]:


hoja = workbook['precipitacion']
hoja["D5"].value


# In[3]:


hoja2 = boletin["DIARIO"]
hoja3 = boletin["ENVIAR"]
hoja2["FU4"].value


# In[4]:


i = hoja2.cell(row=4, column= 177).value
i


# In[15]:

#
#hoja.cell(row = 4, column = 71 - (31 - 1)).value = hoja2.cell(row = 10, column = 169).value
#
#hoja.cell(row = 4, column = 71 - (31 - 1)).value


# In[16]:
columna_prec = 39 #71 #40 + 31 #39


def llenar_datos(hoja1, hoja2, i, columna_prec):
    
    #San Andrés
    hoja1.cell(row = 4, column = columna_prec - (31 - i)).value = hoja2.cell(row = 9, column = 169).value
    #Providencia
    hoja1.cell(row = 5,column = columna_prec - (31 - i)).value = hoja2.cell(row = 10, column = 169).value
    #Santa mayta
    hoja1.cell(row = 6, column = columna_prec - (31 - i)).value = hoja2.cell(row = 11, column = 169).value
    #Cartagena
    hoja1.cell(row = 7, column = columna_prec - (31 - i)).value = hoja2.cell(row = 12, column = 169).value
    #Barranquilla
    hoja1.cell(row = 8, column = columna_prec - (31 - i)).value = hoja2.cell(row = 13, column = 169).value
    #Riohacha 
    hoja1.cell(row = 9, column = columna_prec - (31 - i)).value = hoja2.cell(row = 14, column = 169).value
    #Valledupar
    hoja1.cell(row = 10, column = columna_prec - (31 - i)).value = hoja2.cell(row = 15, column = 169).value
    #Monteria
    hoja1.cell(row = 12, column = columna_prec - (31 - i)).value = hoja2.cell(row = 16, column = 169).value
    #Apartado
    hoja1.cell(row = 13, column = columna_prec - (31 - i)).value = hoja2.cell(row = 20, column = 169).value
    #Barrancabermeja
    hoja1.cell(row = 15, column = columna_prec - (31 - i)).value = hoja2.cell(row = 17, column = 169).value
    #Bucaramanga
    hoja1.cell(row = 16, column = columna_prec - (31 - i)).value = hoja2.cell(row = 18, column = 169).value
    #Cucuta
    hoja1.cell(row = 17, column = columna_prec - (31 - i)).value = hoja2.cell(row = 19, column = 169).value
    #Medellin
    hoja1.cell(row = 18, column = columna_prec - (31 - i)).value = hoja2.cell(row = 21, column = 169).value
    #Rionegro
    hoja1.cell(row = 19, column = columna_prec - (31 - i)).value = hoja2.cell(row = 22, column = 169).value
    #Pereira
    hoja1.cell(row = 21, column = columna_prec - (31 - i)).value = hoja2.cell(row = 23, column = 169).value
    #Armenia
    hoja1.cell(row = 22, column = columna_prec - (31 - i)).value = hoja2.cell(row = 24, column = 169).value
    #Ibague
    hoja1.cell(row = 23, column = columna_prec - (31 - i)).value = hoja2.cell(row = 25, column = 169).value
    #Bogota
    hoja1.cell(row = 24, column = columna_prec - (31 - i)).value = hoja2.cell(row = 26, column = 169).value
    #Cali
    hoja1.cell(row = 27, column = columna_prec - (31 - i)).value = hoja2.cell(row = 27, column = 169).value
    #Neiva
    hoja1.cell(row = 29, column = columna_prec - (31 - i)).value = hoja2.cell(row = 28, column = 169).value
    #Chachagui
    hoja1.cell(row = 30, column = columna_prec - (31 - i)).value = hoja2.cell(row = 29, column = 169).value
    #Ipiales
    hoja1.cell(row = 31, column = columna_prec - (31 - i)).value = hoja2.cell(row = 30, column = 169).value
    #Quibdo
    hoja1.cell(row = 34, column = columna_prec - (31 - i)).value = hoja2.cell(row = 31, column = 169).value
    #Arauca
    hoja1.cell(row = 38, column = columna_prec - (31 - i)).value = hoja2.cell(row = 32, column = 169).value
    #Puerto Carreño
    hoja1.cell(row = 39, column = columna_prec - (31 - i)).value = hoja2.cell(row = 33, column = 169).value
    #Villavicencio
    hoja1.cell(row = 40, column = columna_prec - (31 - i)).value = hoja2.cell(row = 34, column = 169).value
    #Leticia
    hoja1.cell(row = 42, column = columna_prec - (31 - i)).value = hoja2.cell(row = 35, column = 169).value
    
    ############ Areas operativas ########
    
#    j = 0
#    
#    for r in range(1, hoja1.max_row+1):
#        for c in range(1, hoja1.max_column+1):
#            s = str(hoja1.cell(r,c).value)
#            if s == "/" in s: 
#                hoja1.cell(r,c).value = None 
#                #print("row {} col {} : {}".format(r,c,s))
#                j += 1
    
    return workbook.save(filename = r"C:\Users\easalazarm\Documents\IDEAM\IDEAM\Turnos\Diciembre\Tablas\Salidas\precipitacion.xlsx")    


funcion = llenar_datos(hoja, hoja2, i, columna_prec)
funcion


# In[ ]:




